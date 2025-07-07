from datetime import date

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse

from talleres.models import Taller
from asistencia.models import Asistencia, SesionTaller
from asistencia.forms import ReporteAsistenciaForm
from asistencia.utils import generar_reporte_excel


from io import BytesIO
from openpyxl import Workbook
from django.utils.text import slugify

import openpyxl
from openpyxl.styles import Font, Alignment


from .forms import SubirAsistenciaExcelForm
from openpyxl import load_workbook
from cursos.models import Alumno
from django.contrib.auth.decorators import login_required, user_passes_test

from datetime import datetime, date
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import SubirAsistenciaExcelForm
from openpyxl import load_workbook
from cursos.models import Alumno
from .models import Asistencia, Taller

from django.db.models import Count, Q, F
from clientes.models import Colegio
from cursos.models import Curso


import calendar
from datetime import date
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import AsistenciaUpload
from clientes.models import Colegio


import calendar
from datetime import date, timedelta
from django.db.models import Min, Max
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import AsistenciaUpload
from clientes.models import Colegio

from django.db.models import Sum




def can_marcar_asistencia(user):
    return user.is_authenticated and user.rol in ('profesor', 'coordinador', 'administrador')

@login_required
@user_passes_test(can_marcar_asistencia)
def index(request):
    hoy = date.today()

    # Filtrado de talleres según rol
    if request.user.rol == 'profesor':
        talleres_raw = Taller.objects.filter(profesor=request.user)
    elif request.user.rol == 'coordinador':
        talleres_raw = Taller.objects.filter(curso__colegio=request.user.colegio)
    else:  # administrador
        talleres_raw = Taller.objects.all()

    talleres = []
    for t in talleres_raw.select_related('curso'):
        if request.user.rol == 'profesor':
            # Para profesores: mirar SesionTaller
            try:
                sesion = SesionTaller.objects.get(taller=t, fecha=hoy)
                pasada = True
                # Mostrar la etiqueta del estado
                estado = dict(SesionTaller.ESTADOS).get(sesion.estado, '')
            except SesionTaller.DoesNotExist:
                pasada = False
                estado = ''
        else:
            # Para coordinadores/administradores: mirar Asistencia
            pasada = Asistencia.objects.filter(taller=t, fecha=hoy).exists()
            estado = ''

        talleres.append({
            'nombre': t.nombre,
            'curso': t.curso,
            'pk': t.pk,
            'asistencia_pasada': pasada,
            'estado_sesion': estado,
        })

    return render(request, 'asistencia/index.html', {'talleres': talleres})



@login_required
@user_passes_test(can_marcar_asistencia)
def marcar_asistencia(request, taller_id):
    taller = get_object_or_404(Taller, pk=taller_id)
    hoy = date.today()

    try:
        sesion = SesionTaller.objects.get(taller=taller, fecha=hoy)
        if request.method == 'GET':
            messages.info(request, 'Ya pasaste la lista para este taller hoy. Puedes editarla a continuación.')
    except SesionTaller.DoesNotExist:
        sesion = None

    if request.method == 'POST':
        estado = request.POST.get('estado_sesion')
        motivo_extra = request.POST.get('motivo_extra', '').strip()

        sesion, _ = SesionTaller.objects.update_or_create(
            taller=taller,
            fecha=hoy,
            defaults={'estado': estado, 'motivo_extra': motivo_extra}
        )

        if estado == 'R':
            for alumno in taller.alumnos.all():
                key = f'presente_{alumno.pk}'
                est = 'P' if request.POST.get(key) == 'on' else 'F'
                Asistencia.objects.update_or_create(
                    taller=taller,
                    alumno=alumno,
                    fecha=hoy,
                    defaults={'estado': est}
                )
        return redirect('asistencia:index')

    alumnos = taller.alumnos.all().order_by('nombre')

    present_ids = set()
    if sesion and sesion.estado == 'R':
        present_ids = set(
            Asistencia.objects.filter(taller=taller, fecha=hoy, estado='P')
                      .values_list('alumno_id', flat=True)
        )

    return render(request, 'asistencia/marcar.html', {
        'taller': taller,
        'alumnos': alumnos,
        'fecha': hoy,
        'sesion': sesion,
        'present_ids': present_ids,
        'estados': SesionTaller.ESTADOS,
    })


@login_required
@user_passes_test(lambda u: u.rol in ('administrador', 'coordinador'))
def reporte_asistencia(request):
    asistencias = Asistencia.objects.select_related('taller', 'alumno').order_by('-fecha')
    return render(request, 'asistencia/reporte.html', {'asistencias': asistencias})


@login_required
@user_passes_test(lambda u: u.rol in ('administrador', 'coordinador', 'cliente'))
def descargar_reporte_excel(request, taller_id, anio, mes):
    output = generar_reporte_excel(taller_id, mes, anio)
    filename = f"reporte_asistencia_taller_{taller_id}_{anio}_{mes}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@login_required
def seleccionar_reporte(request):
    if request.user.rol == 'profesor':
        talleres = Taller.objects.filter(profesor=request.user)
    elif request.user.rol == 'coordinador':
        talleres = Taller.objects.filter(curso__colegio=request.user.colegio)
    else:
        talleres = Taller.objects.all()

    if request.method == 'POST':
        taller_id = request.POST.get('taller')
        if taller_id:
            return redirect('asistencia:reporte_taller', taller_id=taller_id)

    return render(request, 'asistencia/seleccionar_reporte.html', {'talleres': talleres})


@login_required
@user_passes_test(lambda u: u.rol == 'supervisor')
def supervisar_asistencia(request):
    hoy = date.today()
    talleres_qs = Taller.objects.select_related('curso').all()
    rows = []
    for t in talleres_qs:
        try:
            ses = SesionTaller.objects.get(taller=t, fecha=hoy)
            pasada = True
            estado = ses.estado
        except SesionTaller.DoesNotExist:
            pasada = False
            estado = None
        rows.append({'taller': t, 'pasada': pasada, 'estado': estado})
    return render(request, 'asistencia/supervisar_asistencia.html', {
        'rows': rows,
        'hoy': hoy
    })



@login_required
def reporte_taller(request, taller_id):
    taller = get_object_or_404(Taller, pk=taller_id)
    registros = Asistencia.objects.filter(taller=taller).select_related('alumno').order_by('fecha')
    return render(request, 'asistencia/reporte_taller.html', {
        'taller': taller,
        'registros': registros
    })

@login_required
def descargar_reporte_excel(request, taller_id):
    taller = get_object_or_404(Taller, pk=taller_id)
    registros = Asistencia.objects.filter(taller=taller).select_related('alumno').order_by('fecha', 'alumno__nombre')

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # Encabezado descriptivo
    ws.append(["COLEGIO :", taller.curso.colegio.nombre])
    ws.append(["NOMBRE TALLER", taller.nombre])
    ws.append(["CURSO", str(taller.curso)])
    ws.append([])
    ws.append(["Fecha", "Alumno", "Estado"])

    # Formato
    for col in range(1, 4):
        ws.cell(row=5, column=col).font = Font(bold=True)

    for reg in registros:
        ws.append([
            reg.fecha.strftime("%d/%m/%Y"),
            reg.alumno.nombre,
            reg.get_estado_display()
        ])

    # Descargar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Asistencia_{slugify(taller.nombre)}_{date.today().isoformat()}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


# Vista para subir y procesar el Excel de asistencia
@login_required
def subir_asistencia_excel(request):
    if request.method == 'POST':
        form = SubirAsistenciaExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data['archivo']
            wb = load_workbook(filename=archivo, data_only=True)
            ws = wb.active

            # 1) Leer nombre de taller desde B1
            taller_nombre = (ws['B1'].value or '').strip()
            if not taller_nombre:
                messages.error(request, "El taller no existe.")
                return redirect('asistencia:subir_asistencia_excel')
            try:
                taller = Taller.objects.get(nombre__iexact=taller_nombre)
            except Taller.DoesNotExist:
                messages.error(request, "El taller no existe.")
                return redirect('asistencia:subir_asistencia_excel')

            # 2) Buscar la fila donde col A == "Fecha"
            fecha_cell = None
            for row in range(1, ws.max_row + 1):
                etiqueta = ws.cell(row=row, column=1).value
                if etiqueta and str(etiqueta).strip().lower() == 'fecha':
                    fecha_cell = ws.cell(row=row, column=2).value
                    break
            if fecha_cell is None:
                messages.error(request, 'No se encontró la etiqueta "Fecha".')
                return redirect('asistencia:subir_asistencia_excel')

            # 3) Parsear la fecha
            if isinstance(fecha_cell, datetime):
                fecha = fecha_cell.date()
            elif isinstance(fecha_cell, date):
                fecha = fecha_cell
            else:
                fecha_str = str(fecha_cell).strip()
                try:
                    fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, 'Formato de fecha inválido, usa YYYY-MM-DD.')
                    return redirect('asistencia:subir_asistencia_excel')

            # 4) Procesar lista de alumnos desde la fila 5
            #    Inscritos al curso relacionado con el taller
            inscritos = {
                a.nombre.strip().lower(): a
                for a in Alumno.objects.filter(curso=taller.curso)
            }
            creados = 0
            omitidos = []

            for nombre, marca in ws.iter_rows(min_row=5, min_col=1, max_col=2, values_only=True):
                if not nombre:
                    continue
                if str(marca).strip().upper() == 'X':
                    key = nombre.strip().lower()
                    alumno = inscritos.get(key)
                    if alumno:
                        Asistencia.objects.get_or_create(
                            alumno=alumno,
                            taller=taller,
                            fecha=fecha,
                            defaults={'estado': 'Presente'}
                        )
                        creados += 1
                    else:
                        omitidos.append(nombre.strip())

            # 5) Feedback al usuario
            # 2bis) Crear o actualizar registro de subida
            from .models import AsistenciaUpload
            colegio = request.user.colegio  # asumo que el usuario coordinador tiene FK a Colegio
            AsistenciaUpload.objects.update_or_create(
                colegio=colegio,
                fecha=fecha,
                defaults={'archivo': archivo, 'uploaded_by': request.user}
            )


    else:
        form = SubirAsistenciaExcelForm()

    return render(request, 'asistencia/subir_asistencia.html', {'form': form})


@login_required
@user_passes_test(lambda u: u.rol == 'supervisor')
def supervisar_colegios(request):
    data = []
    for colegio in Colegio.objects.all():
        # total sesiones = suma de numero_sesiones de TODOS sus talleres
        total = Taller.objects.filter(curso__colegio=colegio).aggregate(
            total=Sum('numero_sesiones')
        )['total'] or 0

        # pasadas = contar todas las SesionTaller con estado='R' de esos talleres
        pasadas = SesionTaller.objects.filter(
            taller__curso__colegio=colegio,
            estado='R'
        ).count()

        progreso = int( pasadas * 100 / total ) if total else 0

        data.append({
            'colegio': colegio,
            'total': total,
            'pasadas': pasadas,
            'progreso': progreso,
        })
    return render(request, 'asistencia/supervisar_colegios.html', {'data': data})


@login_required
@user_passes_test(lambda u: u.rol == 'supervisor')
def supervisar_colegios_mes(request):
    from .models import SesionTaller
    from clientes.models import Colegio
    import calendar
    from datetime import date

    hoy = date.today()
    año = int(request.GET.get('year', hoy.year))
    mes = int(request.GET.get('month', hoy.month))

    # 1) Generar lista de días del mes
    _, ultimo_dia = calendar.monthrange(año, mes)
    días = list(range(1, ultimo_dia + 1))

    data = []
    for colegio in Colegio.objects.all():
        # 2) Buscar todas las sesiones programadas de ese mes/año
        sesiones = SesionTaller.objects.filter(
            fecha__year=año,
            fecha__month=mes,
            taller__curso__colegio=colegio
        )
        # 3) Marcar como subidos los días en que hay al menos una SesionTaller
        días_subidos = { sesion.fecha.day for sesion in sesiones }
        data.append({
            'colegio': colegio,
            'días_subidos': días_subidos
        })

    # 4) Meses y años para el selector (igual que antes)
    nombres = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
               'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    meses = [{'num': i+1, 'nombre': nombres[i]} for i in range(12)]

    # Rango fijo ±5 años
    anio_inicio = hoy.year - 5
    anio_fin    = hoy.year + 5
    anios = list(range(anio_inicio, anio_fin + 1))

    # Cálculo mes anterior / siguiente
    if mes == 1:
        prev_month, prev_year = 12, año - 1
    else:
        prev_month, prev_year = mes - 1, año
    if mes == 12:
        next_month, next_year = 1, año + 1
    else:
        next_month, next_year = mes + 1, año

    # Aseguro que los years de las flechas estén en el dropdown
    anios = sorted(set(anios) | {prev_year, next_year})

    return render(request, 'asistencia/supervisar_colegios_mes.html', {
        'año': año,
        'mes': mes,
        'días': días,
        'data': data,
        'meses': meses,
        'anios': anios,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
    })


@login_required
@user_passes_test(lambda u: u.rol == 'profesor')
def pasar_lista(request, taller_id):
    hoy = date.today()
    taller = get_object_or_404(Taller, pk=taller_id)
    alumnos = Alumno.objects.filter(taller=taller)

    # Crear o recuperar la sesión de hoy para este taller
    sesion, _ = SesionTaller.objects.get_or_create(taller=taller, fecha=hoy)

    if request.method == 'POST':
        # 1) Guardar estado de la sesión
        estado = request.POST.get('estado_sesion')
        if estado not in dict(SesionTaller.ESTADOS):
            messages.error(request, "Debes seleccionar un estado válido para la sesión.")
            return redirect('asistencia:marcar', taller_id=taller.pk)

        sesion.estado = estado
        sesion.save()

        # 2) Guardar asistencia de cada alumno
        for alumno in alumnos:
            marca = request.POST.get(f'asistio_{alumno.pk}')
            Asistencia.objects.update_or_create(
                alumno=alumno,
                taller=taller,
                fecha=hoy,
                defaults={'estado': 'P' if marca == 'on' else 'F'}
            )

        messages.success(request, "Asistencia y estado de sesión guardados correctamente.")
        return redirect('asistencia:marcar', taller_id=taller.pk)

    # GET: renderizar el formulario precargando el estado actual
    return render(request, 'asistencia/pasar_lista.html', {
        'taller': taller,
        'alumnos': alumnos,
        'sesion': sesion,
    })