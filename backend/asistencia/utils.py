import calendar
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font

from .models import Asistencia


def generar_reporte_excel(taller_id, mes, anio):
    # Crear workbook y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Asistencia"

    # Encabezado mes y año
    mes_nombre = calendar.month_name[mes]
    titulo = f"Mes: {mes_nombre} - Año: {anio}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=40)
    celda_titulo = ws.cell(row=1, column=1)
    celda_titulo.value = titulo
    celda_titulo.font = Font(size=14, bold=True)
    celda_titulo.alignment = Alignment(horizontal="center")

    # Filas iniciales para nombres y cursos
    ws.cell(row=3, column=1, value="N°")
    ws.cell(row=3, column=2, value="Nombre Alumno(a)")
    ws.cell(row=3, column=3, value="Curso")

    # Obtener todos los días del mes (filtrar sólo los días donde hay clases)
    dias_mes = [day for day in range(1, calendar.monthrange(anio, mes)[1] + 1)]

    # Colocar encabezado de días (desde col 4 en adelante)
    col_dia_inicio = 4
    for idx, dia in enumerate(dias_mes):
        celda = ws.cell(row=2, column=col_dia_inicio + idx)
        fecha = datetime(anio, mes, dia)
        celda.value = fecha.strftime("%a %d/%m")
        celda.alignment = Alignment(horizontal="center")

    # Columna Total Clases
    total_col = col_dia_inicio + len(dias_mes)
    ws.cell(row=2, column=total_col, value="Total Clases")

    # Obtener asistencias del taller para el mes/año
    asistencias = Asistencia.objects.filter(
        taller_id=taller_id,
        fecha__year=anio,
        fecha__month=mes
    ).select_related('alumno', 'taller', 'taller__curso')

    # Crear dict para acceder rápido: {(alumno_id, fecha): estado}
    asistencia_dict = {}
    for a in asistencias:
        asistencia_dict[(a.alumno_id, a.fecha.day)] = a.estado

    # Obtener lista ordenada de alumnos del taller
    alumnos = asistencias.values_list('alumno_id', 'alumno__nombre', 'taller__curso__grado').distinct().order_by('alumno__nombre')

    # Si no hay asistencias, intentamos sacar alumnos desde el taller para listar (sin estados)
    if not alumnos:
        from talleres.models import Taller
        taller = Taller.objects.get(id=taller_id)
        alumnos = [(alumno.pk, alumno.nombre, alumno.curso.grado) for alumno in taller.alumnos.all().order_by('nombre')]

    # Poner datos fila por fila
    fila = 4
    for idx, alumno in enumerate(alumnos, start=1):
        if isinstance(alumno, tuple):
            alumno_id, nombre, curso = alumno
        else:
            # caso diferente, solo id
            alumno_id, nombre, curso = alumno, "", ""
        ws.cell(row=fila, column=1, value=idx)  # nro
        ws.cell(row=fila, column=2, value=nombre)
        ws.cell(row=fila, column=3, value=str(curso))

        total_clases = 0
        for i, dia in enumerate(dias_mes):
            estado = asistencia_dict.get((alumno_id, dia), "")
            celda = ws.cell(row=fila, column=col_dia_inicio + i)
            celda.value = estado
            celda.alignment = Alignment(horizontal="center")
            if estado not in ("SUSPENDIDO", "FERIADO", ""):
                total_clases += 1

        ws.cell(row=fila, column=total_col, value=total_clases)
        fila += 1

    # Ajustar ancho columnas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    for col in range(col_dia_inicio, total_col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 8

    # Guardar en buffer bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
